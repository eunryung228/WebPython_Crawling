from openpyxl import Workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import time


class ReviewAnalyzing:
    def __init__(self, fileName, sheetName):
        self.fileName=fileName
        self.sheetName=sheetName
        self.urlList=[]

    def MakeUrlList(self):
        with open(self.fileName, 'r', encoding="utf8") as fileRead:
            for lineContent in fileRead:
                if (lineContent != '\n'):  # 줄넘김 문자는 제외
                    self.urlList.append(lineContent[lineContent.find("https://"):].strip('\n'))
                    # 게임 이름이나 회사 중간에 ','가 들어있는 경우 때문에 split(',')으로는 구분 불가능

    def GetReview(self):
        count = 0
        reviewNum = 1
        path_1 = '//*[@id="fcxH9b"]/div[4]/c-wiz/div/div[2]/div/div[1]/div/div/div[1]/div[2]/div/div['
        path_review = ']/div/div[2]/div[2]/span[1]'
        path_allReview = ']/div/div[2]/div[2]/span[2]'

        while count < len(self.urlList):
            driver.get(self.urlList[count] + '&showAllReviews=true')

            for i in range(5):
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(3)  # 스크롤을 내려 로딩한 값까지 긁어오기 위해서 3초 대기

            for i in range(1, 201):
                try:
                    driver.find_element_by_xpath(path_1 + str(i) + path_review + "/div/button")
                    display = driver.find_element_by_xpath(path_1 + str(i) + path_allReview)
                    print(display.get_attribute('textContent'))
                    ws['A' + str(reviewNum)] = display.get_attribute('textContent')
                    reviewNum += 1
                except NoSuchElementException:
                    try:
                        display = driver.find_element_by_xpath(path_1 + str(i) + path_review)
                        print(display.text)
                        ws['A' + str(reviewNum)] = display.text
                        reviewNum += 1
                    except NoSuchElementException:
                        break
            print("\nReview: ", reviewNum)
            count += 1

driver=webdriver.Chrome("/Users/sayqu/Desktop/은령/과제/3-1/웹파/TermProject/chromedriver")
wb=Workbook()

upperReview=ReviewAnalyzing("Game_Rank_upperList.csv", 'UpperList')
upperReview.MakeUrlList()
ws=wb.create_sheet(upperReview.sheetName, 0)
upperReview.GetReview()

lowerReview=ReviewAnalyzing("Game_Rank_lowerList.csv", 'LowerList')
lowerReview.MakeUrlList()
ws=wb.create_sheet(lowerReview.sheetName, 0)
lowerReview.GetReview()

wb.save('Review.xlsx')